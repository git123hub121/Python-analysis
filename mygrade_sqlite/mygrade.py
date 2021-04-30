import sqlite3
import openpyxl

#创建一个数据库文件
sql = '''
    create table mygrade
        (
        Id integer primary key autoincrement,
        Semester text,
        C_number text,
        Grade numeric,
        Credit numeric,
        Thours numeric,
        Gpa numeric, 
        Assessment text,
        C_attributes text,  
        Curriculum text
        )
    '''
conn = sqlite3.connect('grade.db')#创建数据库文件
cursor = conn.cursor()#获取游标
cursor.execute(sql)#执行数据库操作
conn.commit()#提交数据库操作
cursor.close()
conn.close()

lists = sqlite3.connect('grade.db')
cur = lists.cursor()
listinsheet = openpyxl.load_workbook('grade.xlsx')
datalist = listinsheet.active
dtat_sql = '''INSERT INTO mygrade(Id,Semester,C_number,Grade,Credit,Thours,Gpa,Assessment,C_attributes,Curriculum)
VALUES(?,?,?,?,?,?,?,?,?,?)'''
for row in datalist.iter_rows(min_row=2,max_col=10,max_row=datalist.max_row):
    cargo = [cell.value for cell in row]
    cur.execute(dtat_sql,cargo)
lists.commit()
lists.close()
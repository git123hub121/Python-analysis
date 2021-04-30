from openpyxl import load_workbook
workbook = load_workbook(filename="test.xlsx")
#workbook.sheetnames
sheet = workbook.active
print(sheet)
# print(sheet.dimensions)
#
# cell1 = sheet.cell(row = 1,column = 1)
# cell2 = sheet.cell(row = 11,column = 3)
# print(cell1.value, cell2.value)
#
# cell3 = sheet["A2"]
# cell4 = sheet["B11"]
# print(cell3.value,cell4.value)
#
# cell = sheet["A1:C2"]
# print(cell)
# for i in cell:
#     for j in i:
#         print(j.value)

# sheet["A"] --- 获取 A 列的数据
# sheet["A:C"] --- 获取 A,B,C 三列的数据
# sheet[5] --- 只获取第 5 行的数据
# for i in sheet.rows:
#     print(i)
# sheet["A1"]="哈罗"
# workbook.save(filename="test.xlsx")
# data = [
#     ["唐僧1","nan","180cm"],
#     ["唐僧2","nan","180cm"],
#     ["唐僧3","nan","180cm"],
#     ["唐僧4","nan","180cm"],
# ]
# for row in data:
#     sheet.append(row)
# workbook.save(filename="test.xlsx")
#在 python 中使用 excel 函数公式(很有用)
# sheet.insert_cols(idx=4,amount=2)
# sheet.insert_rows(idx=5,amount=4)
# sheet.move_range("C1:D4",rows=2,cols=-1)
workbook.create_sheet("我是一个新的 sheet")
print(workbook.sheetnames)
workbook.save(filename = "test.xlsx")

